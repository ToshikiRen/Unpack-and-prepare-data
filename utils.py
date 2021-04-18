import zipfile
import re, os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import patoolib
import shutil
import codecs
import mosspy
import time
import subprocess
# sheet_names = ["321AA", "322AA", "323AA", "324AA", "Restantieri (AA)", "321AB",
#                "322AB", "323AB", "324AB", "321AC", "322AC", "323AC", "324AC"]
sheet_names = ["312AB"]

class Friend():

    def __init__(self, data):
        [self.name, self.percent] = return_name_percent(data)
        self.name = addSpace(self.name)
    
    def get_name(self):
        return self.name

    def get_percent(self):
        return self.percent

class Friends:

    def __init__(self, first, second, lines):
        self.first = Friend(first)
        self.second = Friend(second)
        self.lines = lines

    def __str__(self):
        first = self.first.get_name()
        second = self.second.get_name()
        lines = self.lines

        return first + ' si ' + second + ' au ' + str(int(lines)) + ' de linii identice!'

    def get_name_first(self):
        return self.first.get_name()

    def get_name_second(self):
        return self.second.get_name()
    
    def get_percent_first(self):
        return self.first.get_percent()

    def get_percent_second(self):
        return self.second.get_percent()

    def get_lines(self):
        return self.lines

    def get_mean_percent(self):
        return (self.first.get_percent() + self.second.get_percent())/2

# Adds one space between firstname and lastname
def addSpace(nameString):
    
    nameList = []
    nameList.append(nameString[0])
    for i in range(1, len(nameString)):
        if nameString[i - 1].islower() and nameString[i].isupper():
            nameList.append(" ")
        nameList.append(nameString[i])
    return "".join(nameList)


# Returns data from moss sheet while removing the path
def return_data(active_sheet, desired_collums, substringData = 'C:/Users/leona/OneDrive/Desktop/Moss/'):
    
    firstCol = active_sheet[desired_collums[0]]
    data = [ [] for j in range(len(firstCol)) if firstCol[j].value != None ]
    
    for col in desired_collums:
    
        for row in range(0,len(active_sheet[col])):
            cell = active_sheet[col][row].value
            try:
                if cell != None:
                    data[row].append(remove_substr(cell, substringData))
            except:
                pass
    return data

#Retuns only the ones form the provided file
def return_my_sudents(students_grades, student_list, output_file):
    
    file = open(student_list, mode="r", encoding="utf-8")
    our_students = [student.replace('\n', "") for student in file]
    
    file_output = open(output_file,  mode = 'w', encoding = "utf-8")
    
    for student in students_grades:
    
        if student[0] in our_students:
            file_output.write(student[0] + ' - ' + str(student[1]) + '\n');
    
    


# Returns the data from the moss sheet wraped in a class
def return_class_data(active_sheet, desired_collums, substringData = 'C:/Users/leona/OneDrive/Desktop/Moss/'):
    
    data = return_data(active_sheet, desired_collums, substringData)
    returned_data = []
    for row in range(len(data)):
        returned_data.append(Friends(data[row][0], data[row][1], data[row][2]))
    
    return returned_data

# String removal with try to avoid type errors
def remove_substr(stringData, substringData):
    try:
       return_data = stringData.replace(substringData, '')
    except:
        return stringData
    return return_data


#Returns the fullname and similarity percent
def return_name_percent(stringData):

    name = "".join([c for c in stringData if not c.isdigit() and c not in "()%/"])
    percent = float(re.sub('[^0-9]+', '', stringData))
    name = name[:-1]
    return [name, percent]

# Removes submission data
def remove_submission_data(stringData, moss = True):
    stringData = remove_substr(stringData, '_assignsubmission_file_')
    stringData = "".join([y for y in stringData if not y.isdigit() and y not in "_()%/"])

    if moss:
        return stringData[:-1]
    return stringData
    
# Renames all the files from moodle in order to keep only the fullname
# of the student
def rename_all(folderName):

    initialFolder = os.getcwd()
    os.chdir(folderName)
    submissionList = os.listdir()

    for friend in submissionList:
        os.rename(friend, remove_submission_data(friend, False))

    os.chdir(initialFolder)

# Color one row
def setRowColor(sheet, index, color):
    for colorRow in sheet.iter_rows(min_row = index, max_row = index, min_col = 1):
        for cell in colorRow:
            cell.fill = PatternFill(start_color = color,  fill_type = "solid")

# Creates an .xlsx where all the naughty ones have 0 and are colored red
def above_threshold(catalogName, class_data, threshold, gradeCol, make_red = True):
    
    catalog = load_workbook(filename = catalogName)
    naughtySet = set()
    
    # Find all the naughty ones
    for friends in class_data:

        if friends.get_mean_percent() >= threshold:
            naughtySet.add(friends.get_name_first())
            naughtySet.add(friends.get_name_second())
    
    # Mark all naughty ones 
    for fiend in naughtySet:
        executed = False
        for sheet_name in sheet_names:
            sheet = catalog[sheet_name]
            
            for row in sheet.iter_rows(min_row = 2, min_col = 1):
                name = row[1].value

                if name == None:
                    break
                index = int(row[0].value) + 1
                
                # Mark the naughty one 
                if name == fiend:
                    executed = True
                    if make_red:
                        setRowColor(sheet, index, 'ff0000')
                    row[gradeCol - 1].value = 0
                    break
                    
            if executed:
                break
                
    now = datetime.now()
    current_time = now.strftime("%H-%M-%S")
    catalogName = catalogName.replace('.xlsx', '') + str(current_time)+ '.xlsx'
    catalog.save(catalogName)

    return naughtySet

# Unzip files in folder and rename unzipfile
def unzip_files(folder, subFolderName = "src"):
    
    naughtyOnes = open("naughtyOnes.txt", mode = "w", encoding = 'UTF-8')
    
    
    initialFolder = os.getcwd()
    os.chdir(folder)
    folder_list = os.listdir()
    root_folder = os.path.join(initialFolder, folder)

    for attempt in folder_list:
        # Path to attemt 
        attempt_folder = os.path.join(root_folder, attempt)
        os.chdir(attempt_folder)
        # Get the first zip-file
        zip_file = os.listdir()
        
        try:
            zip_file = zip_file[0]
            path_to_zip_file = os.path.join(attempt_folder, zip_file)
        except:
            naughtyOnes.write(str(attempt.encode('UTF-8')) + ' n-a incarcat o arhiva\n')
            continue
        try:
            # Unzip file
            patoolib.extract_archive(zip_file, outdir = attempt_folder)
            os.remove(zip_file)
            files = os.listdir()
            
            # Zip contains only the src folder
            if len(files) == 1:
                file = files[0]
                filePath = os.path.join(attempt_folder, file)
                if os.path.isdir(filePath):
                    os.rename(file, subFolderName)
            else:
                # Zip has many folders
                naughtyOnes.write(str(attempt.encode('UTF-8')) + ' n-a incarcat src-ul\n')
                os.mkdir(subFolderName)
                
                for file in files:
                   
                    destination = os.path.join(attempt_folder, subFolderName)
                    source = os.path.join(attempt_folder, file)
                    
                    shutil.move(source , destination)

        except:
            naughtyOnes.write(str(attempt) + ' n-a incarcat o arhiva ok\n')
    naughtyOnes.close()
    os.chdir(initialFolder)
   
# Place all students form student_list in the destination_folder
def get_students(source_folder, student_list, destination_folder, in_file = True):

    current_directory = os.getcwd()
    
    file = open(student_list, mode="r", encoding="utf-8")
    our_students = [student.replace('\n', "") for student in file]
    
    # Create the destination_folder or overwrite it
    try:
        os.mkdir(destination_folder)
    except:
        shutil.rmtree(os.path.join(current_directory, destination_folder))
        os.mkdir(os.path.join(current_directory, destination_folder))
    
    # Go in the source folder
    os.chdir(source_folder)
    
    # Get file names and setup src and dst path
    students_submissions = os.listdir()
    source_path = os.path.join(current_directory, source_folder)
    destination_path = os.path.join(current_directory, destination_folder)
    
    # Loop through the submissions
    for student in students_submissions:
        
        if student in our_students and in_file:
            
            source_path_student = os.path.join(source_path, student)
            destination_path_student = os.path.join(destination_path, student)
            try:
                os.mkdir(destination_path_student)
            except:
                shutil.rmtree(destination_path_student)
                os.mkdir(destination_path_student)
                
            copy_all(source_path_student, destination_path_student)
        elif student not in our_students and not in_file:
            
            source_path_student = os.path.join(source_path, student)
            destination_path_student = os.path.join(destination_path, student)
            try:
                os.mkdir(destination_path_student)
            except:
                shutil.rmtree(destination_path_student)
                os.mkdir(destination_path_student)
                
            copy_all(source_path_student, destination_path_student)

# Check if a file has one of the given extensions
def file_ending_in(file, endings):
    
    for end in endings:
        if file.endswith(end):
            return True
    return False

# Remove all files from the submissions except the arhives
def remove_all_but_arhives(submission_folder):
    
    endings = ['.zip', '.rar', '.taz', '.z']
    current_directory = os.getcwd()
    os.chdir(submission_folder)
    submission_directory = os.getcwd()
    students_submissions = os.listdir()
    
    for student in students_submissions:
        student_path = os.path.join(submission_directory, student)
        os.chdir(student_path)
        
        files = os.listdir()
        for file in files:
            if not file_ending_in(file, endings):
                path = os.path.join(student_path, file)
                if os.path.isdir(file):
                    shutil.rmtree(path)
                else:
                    os.remove(path)
                
def move(src_path, file, moss_dest, endings = []):
    file_path = os.path.join(src_path, file)
    if len(endings):
        if file_ending_in(file, endings):
            shutil.move(file_path, moss_dest)
    else:
        try:
            shutil.move(file_path, moss_dest)
        except:
            pass
            
# move the files with a given ending while keeping the directory structure
def move_all(src, dest, endings = []):
    
    os.chdir(src)
    files = os.listdir(src)
    try:
        os.mkdir(dest)
    except:
        pass
    for file in files:
        if os.path.isdir(file):
            try:
                os.mkdir(os.path.join(dest, file))
            except:
                shutil.rmtree(os.path.join(dest, file))
                os.mkdir(os.path.join(dest, file))
            
            move_all(os.path.join(src, file), os.path.join(dest, file), endings)
           
            os.chdir(src)
        else:
            move(src, file, dest, endings)

# copy one file with one of the given endings
def copy(src_path, file, moss_dest, endings = []):
    file_path = os.path.join(src_path, file)
    if len(endings):
        if file_ending_in(file, endings):
            shutil.copy(file_path, moss_dest)
    else:
        try:
            shutil.copy(file_path, moss_dest)
        except:
            pass
            
# copy the files with a given ending while keeping the directory structure
def copy_all(src, dest, endings = []):
    
    os.chdir(src)
    files = os.listdir(src)
    try:
        os.mkdir(dest)
    except:
        pass
    for file in files:
        if os.path.isdir(file):
            try:
                os.mkdir(os.path.join(dest, file))
            except:
                shutil.rmtree(os.path.join(dest, file))
                os.mkdir(os.path.join(dest, file))
            
            copy_all(os.path.join(src, file), os.path.join(dest, file), endings)
            os.chdir(src)
        else:
            copy(src, file, dest, endings)

# copy all files with one of the given endings to dest folder
def copy_all_TO(src, dest, endings = []):
    
    os.chdir(src)
    files = os.listdir(src)
    for file in files:
        if os.path.isdir(file):            
            copy_all_TO(os.path.join(src, file), dest, endings)
            os.chdir(src)
        else:
            copy(src, file, dest, endings)
    

# Prepare data for moss by moving all the files with one of the given
# endings to one folder for each student  
def prepare_moss(submission_folder, sub_folder_name = 'src',
         endings = ['.h', '.c'], keep_structure = False, folder_ending = ' for Moss'):

    current_directory = os.getcwd()
    
    moss_file_dest = submission_folder + folder_ending
    
        
    try:
        os.mkdir(moss_file_dest)
    except:
        shutil.rmtree(moss_file_dest)
        os.mkdir(moss_file_dest)
    
    moss_file_dest = os.path.join(current_directory, moss_file_dest)
    
    os.chdir(submission_folder)
    submission_directory = os.getcwd()
    students_submissions = os.listdir()
   
    for student in students_submissions:
        student_path = os.path.join(submission_directory, student)
        moss_dest = os.path.join(moss_file_dest, student)
        try:
            os.mkdir(moss_dest)
        except:
            shutil.rmtree(moss_dest)
            os.mkdir(moss_dest)
        
        if len(sub_folder_name):
            src_path = os.path.join(student_path, sub_folder_name);
        else:
            src_path = student_path
            
        try:
            os.mkdir(moss_dest)
        except:
            shutil.rmtree(moss_dest)
            os.mkdir(moss_dest)
        
        if not keep_structure:
            copy_all_TO(src_path, moss_dest, endings)
        else:
            copy_all(src_path, moss_dest, endings)
        
        
        
def add_missing(submission_folder, path_needed, sub_folder_name = 'src'):

    current_directory = os.getcwd()

    print(path_needed)

    os.chdir(current_directory)
    submission_directory = os.path.join(current_directory, submission_folder)
    os.chdir(submission_directory)
    
    print(submission_directory)
    
    students_submissions = os.listdir()
    
    for student in students_submissions:
        
        student_path = os.path.join(submission_directory, student)
        
        destination = os.path.join(student_path, sub_folder_name)
        move_all(student_path, os.path.join(student_path, destination))
        copy_all(path_needed, os.path.join(student_path, destination))
        

def check_homework(submission_directory, sandbox, command = 'make test',
        output_file = 'students_grades.txt', to_remove = 'src'):

    root = os.getcwd()
    output = codecs.open(output_file, 'a', "utf-8")
    output.close();
    os.chdir(submission_directory)
    src = os.getcwd()
    dest = os.path.join(root, sandbox)
    to_remove = os.path.join(dest, to_remove)
    
    # Get students submission list
    students = os.listdir()
    for student in students:
        os.chdir(root)
        output = codecs.open(output_file, 'a', "utf-8")
        os.chdir(src)
        student_src = os.path.join(src, student)
        # Delete the last student src
        os.system('rm -rf ' + to_remove)
        # Copy current student src
        copy_all(student_src, dest)

        os.chdir(dest)
        
        #result = os.system(command)
        #if result == 0: 
        #    output.write(student + '\n')        
        
        try:
            result = subprocess.run(command, timeout = 60)
            if result.returncode == 0: 
                output.write(student + '\n')
        except:
            pass
       
        output.close()
        
        
def remove_files(list_rem = ['312AB', '312AB-for-grading', '312AB-for-Moss',
        'left_to_grade', 'submissions', 'data', 'naughtyOnes'], remake = ['submissions', 'data']):
    
    
    
    root = os.getcwd()
    for file in list_rem:
        try:
            shutil.rmtree(os.path.join(root, file))
        except:
            print('File or Folder not found!')
    
    

    for file in remake:
        try:
            os.mkdir(file)
        except:
            pass
	
	
	
def send_moss_dir_mode(submission, basefile, language):

    userid = 78325144
    
    root = os.getcwd()

    moss = mosspy.Moss(userid, language)
    moss.setDirectoryMode(True)
    
    os.chdir(basefile)
    for base in os.listdir():
        path = os.path.join(os.getcwd(), base)
        if os.path.isfile(base):
            moss.addBaseFile(path)

    os.chdir(root)
    os.chdir(submission)

    submissions = os.getcwd()
    students = os.listdir()

    for student in students:
    
        student_submission = os.path.join(submissions, student)
        os.chdir(student_submission)
        for s_file in os.listdir():
            if os.path.isfile(s_file):
                student_file = os.path.join(student_submission, s_file)    
                moss.addFile(student_file)
        
        os.chdir(submissions)
   
   
    url = moss.send()
    print()
    moss.saveWebPage(url, "report.html")