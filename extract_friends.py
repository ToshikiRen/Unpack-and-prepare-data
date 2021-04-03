from openpyxl import load_workbook
from utils import return_class_data, return_name_percent, return_data, prepare_moss, add_missing
from utils import above_threshold, rename_all,unzip_files, get_students, remove_all_but_arhives, return_my_sudents
from utils import check_homework
import os
import sys

# Path to files sent to MOSS
# remove_text = 'C:/Users/leona/OneDrive/Desktop/Newfolder/'
# Load moss data from file
# sheet = load_workbook(filename = 'sheet.xlsx')
# sheet = sheet.active
# data = return_class_data(sheet, ['A', 'B', 'C'], remove_text)

# needRename = input("Redenumire fisiere = 1. Raspuns = ")
# if needRename == "1":
#     folder = input("Nume folder:")
#     rename_all(folder)

# needUnzip = input("Extractare fisiere = 1. Raspuns = ")
# if needUnzip == "1":
#     folder = input("Nume folder:")
#     unzip_files(folder)

# FIRST PARAMETER = function:
#       0 -> extragere studenti grupa x
#       1 -> redenumire fisiere folder x
#       2 -> dezarhivare fisiere din folder x


argvs = sys.argv

run_folder = os.getcwd()

if argvs[1] == '0':
    source_folder = argvs[2]
    student_list = argvs[4]
    destination_folder = argvs[3]
    get_students(source_folder, student_list, destination_folder)

if argvs[1] == '1':
    folder = argvs[2]
    rename_all(folder)

if argvs[1] == '2':
    folder = argvs[2]
    unzip_files(folder)
    
if argvs[1] == '3':
    remove_text = 'C:/Desktop/SDA/312AB/'
    # Load moss data from file
    sheet = load_workbook(filename = 'sheet.xlsx')
    sheet = sheet.active
    data  = return_class_data(sheet, ['A', 'B', 'C'], remove_text)
    naughty_set = above_threshold('note.xlsx', data, 15, 3, make_red = False)
    file = open('312AB_Copiat.txt', mode = 'w', encoding = 'UTF-8')
    for x in naughty_set:
        file.write(str(x) + '\n')
        
if argvs[1] == '4':
    submission_folder = argvs[2]
    remove_all_but_arhives(submission_folder)
 
if argvs[1] == '5':
    submission_folder = argvs[2]
    sub_folder_name = argvs[3]
    prepare_moss(submission_folder, sub_folder_name)

if argvs[1] == '6':
    submission_folder = argvs[2]
    prepare_moss(submission_folder, '', ['.c'], False, '-for-grading')
   
    path_to_data = os.path.join(run_folder, 'data')
   
    os.chdir(run_folder)
    add_missing(submission_folder + '-for-grading', path_to_data)
    
    os.chdir(run_folder)
    prepare_moss(submission_folder, '', ['.c'], False, '-for-Moss')

if argvs[1] == '7':
    sheet = load_workbook(filename = argvs[2])
    sheet = sheet.active
    data = return_data(sheet, ['A', 'B'], '')
    return_my_sudents(data, argvs[3], argvs[4])

if argvs[1] == '8':
    submission_folder = argvs[2]
    sandbox = argvs[3]
    output_file = argvs[4]
    check_homework(submission_folder, sandbox, output_file)
# Set the grade collum
# gradeCol = 11
# Make all naughty ones red in the .xlsx
# num = above_threshold('Catalog.xlsx', data, 0, gradeCol)
# print(num)


